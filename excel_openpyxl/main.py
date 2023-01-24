# coding: utf-8
## @author Francesco Petracci
## @date 24/01/22
## Goal of this simple script is to highlight some basic openpyxl functions to read and write two lists

# openpyxl is a Python library to read/write Excel 2010 xlsx/xlsm/xltx/xltm files.
# Official documentation: https://openpyxl.readthedocs.io/en/stable/
import openpyxl as oxl

####
# read brand_product
wb_brand_product = oxl.load_workbook("list_brand_product.xlsx")

ws_brand_product = wb_brand_product[wb_brand_product.sheetnames[0]]

brands      = ws_brand_product['A']
products    = ws_brand_product['B']

brands_products = []
for i in range(len(brands)):
    if i == 0:
        pass # skip header row
    else:
        brands_products.append([brands[i].value, products[i].value])
print(brands_products)
print("%"*20)

####
# list product-brand
wb_supermarket_brand = oxl.load_workbook("list_supermarket_brand.xlsx")


ws_supermarket_brand = wb_supermarket_brand[wb_supermarket_brand.sheetnames[0]]

supermarkets = ws_supermarket_brand['A']
brands2      = ws_supermarket_brand['B']

supermarkets_brands2 = []
for i in range(len(supermarkets)):
    if i == 0:
        pass # skip header row
    else:
        supermarkets_brands2.append([supermarkets[i].value, brands2[i].value])
print(supermarkets_brands2)
print("%"*20)

# create a list that links supermarket and products
supermarkets_products = []
# products_unique = list(set(products)) # this eliminates the possible duplicates
for el in supermarkets_brands2:
    supermarket = el[0]
    brand       = el[1]

    for el2 in brands_products:
        brand2  = el2[0]
        product = el2[1]
        if brand == brand2:
            supermarkets_products.append([supermarket, product])

print(supermarkets_products)
print("%"*20)


###
# build a new workbook
output_wb   = oxl.Workbook()
output_path = "output.xlsx" 

ws = output_wb.active
ws.title = "Sheet1"

#supermarkets_products = [["coop", "spaghetti"], ["esselunga", "fusilli"]]

for row in range(len(supermarkets_products) + 1):
    row_xl = row + 1
    if row_xl == 1:
        ws.cell(column=1, row=row_xl).value = "supermarket"
        ws.cell(column=2, row=row_xl).value = "product"
    else: 
        i = row - 1
        ws.cell(column=1, row=row_xl).value = supermarkets_products[i][0]
        ws.cell(column=2, row=row_xl).value = supermarkets_products[i][1]

output_wb.save(filename = output_path)
print("Saved: %s" %output_path)
print("All Done!")


# ws_supermarket_brand.cell(column=1, row=1).value = 'testtest'
# print( ws_supermarket_brand.cell(column=1, row=1).value )