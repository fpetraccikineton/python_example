from copy import copy
import openpyxl
from openpyxl.utils import get_column_letter

def save_workbook_excel_file(workbook, filename: str):
    """Tries to save created data to excel file"""
    try:
        workbook.save(filename)
    except PermissionError:
        print("Error: No permission to save file.")

def invert_row_column(filename: str, filename_out: str):
    """
    Main loop to invert rows and column
    """
    workbook = openpyxl.load_workbook(filename)
    sheet_names = workbook.sheetnames
    sheet = workbook[sheet_names[0]]
    workbook.create_sheet(index=0, title='tmp_sheet')
    tmp_sheet = workbook['tmp_sheet']

    data = []
    for row in sheet:
        cells = []
        for cell in row:
            cells.append(cell)
        data.append(cells)

    for x in range(0, len(data)):
        for y in range(0, len(data[x])):
            column = get_column_letter(x + 1)
            row = str(y + 1)
            tmp_sheet[column + row] = copy(data[x][y].value)

    sheet_name = sheet.title
    del workbook[sheet_name]
    tmp_sheet.title = sheet_name
    save_workbook_excel_file(workbook, filename_out)